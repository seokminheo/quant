import pandas as pd
import pandas_datareader.data as web
import datetime
import backtrader as bt
import numpy as np
import matplotlib.pyplot as plt
import pyfolio as pfd
import quantstats
plt.rcParams["figure.figsize"] = (10, 6) # (w, h)
%matplotlib inline
import sys
from scipy.stats import rankdata
from scipy.stats import stats
from scipy.optimize import minimize
import math
from openpyxl import workbook 
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import numba as nb
import seaborn
from sklearn.preprocessing import StandardScaler, MinMaxScaler
from sklearn.model_selection import train_test_split
from sklearn.utils import shuffle
import tensorflow as tf
import tensorflow_probability as tfp
import matplotlib.pyplot as plt
from tensorflow.keras.layers import Input, LSTM, Dense, Activation
from tensorflow.keras.models import Model
from tensorflow.keras.optimizers import Adam
from sklearn.utils import shuffle
from tensorflow.keras.layers import Add, Dense, Dropout, Embedding, Layer, LayerNormalization, Multiply, Permute, Reshape  # pylint: disable=import-error
import math
from tensorflow.keras.models import load_model


import ta
import datetime as dt
import pandas_datareader.data as web
from time import sleep
import pickle
from keras.callbacks import EarlyStopping




data_path = 'c:/users/user/desktop/Python/기초 자산배분 전략/regular/'
df_asset = pd.read_excel(data_path +'deep learning.xlsx', sheet_name = 0,
                         index_col = 0, parse_dates = True)

# plot the recent history
log_returns = np.log(df_asset.loc[df_asset.loc['2023-01'].index[-1]:,:]).diff()
m = log_returns.cumsum()
fig, ax = plt.subplots(figsize=(15,8), dpi = 400)
ax.plot(m)
fig.legend(m.columns)
ax.set_xlim([m.index[0], m.index[-1]])
ax.grid()
ax.set_title('Asset Universe - Latest Month')


# render input asset price data into return data
price_df = df_asset.iloc[:,0:5]
return_df = price_df.pct_change().dropna(axis=0)

# render output style price data into return data(natural log)



@nb.jit
def objective(w_o, *args):
    rtn_o = args[0]
    cm_o = args[1]
    gamma_o = args[2]
    reg_o = args[3]

    # annual return
    port_rtn = np.dot(w_o, rtn_o) * 252
    
    # annual volatility
    port_var = np.dot(w_o, np.dot(cm_o, w_o.T)) * 252
    
    # sharpe ratio
    SR = port_rtn - gamma_o * port_var - reg_o * np.sum(w_o ** 2)
    # SR = port_rtn / port_var
    # minimize -> insert (-)
    return -SR

# Long only, No leverage, Weights sum up to 1
@nb.jit
def constraint1(w_o):
    return np.sum(w_o) - 1.0

# Portfolio Optimization
# gamma : L1 constraint
# reg : L2 constraint
def opt_portfolio(rv, W0, gamma=0.1, reg=0.1):
    R = np.mean(rv, axis=0)    
    C = np.cov(rv.T)           
    
    # Calculate the optimal weights
    # weight range = 1% ~ 50%
    bnds = ((0.01, 0.5),) * len(W0)
    cons = {'type' : 'eq', 'fun' : constraint1}
    p = minimize(fun=objective, x0=W0, args=(R, C, gamma, reg), 
                          method='SLSQP', bounds=bnds, constraints=cons,
                          options={'ftol': 1e-20, 'maxiter': 2000})
    
    return p.x

# Hyperparameters
N_STOCKS = return_df.shape[1]
W0 = np.ones(N_STOCKS) / N_STOCKS

# Chose appropriate level
gamma = 3.0
reg = 5.0     # regularization constant (to prevent overfitting)
# w = opt_portfolio(np.array(return_df), W0, gamma, reg)

# 252 days rolling 
weight_df = pd.DataFrame(columns=['US equities','7-10Y UST','10-20Y UST','Gold','Commodities'])

days = 252

for i in range(len(return_df)-days+1):
    
    print(i)
    temp_return_df = return_df.iloc[i:(i+days), :]
    date_index = str(temp_return_df.iloc[days-1, :].name)[0:10]
    temp = pd.DataFrame(opt_portfolio(np.array(temp_return_df.iloc[0:days, :]), 
        W0, gamma, reg), index=['US equities','7-10Y UST',
                                '10-20Y UST','Gold','Commodities'], columns=[date_index]).T
    weight_df = pd.concat([weight_df, temp])

weight_df.plot()


def ewma_df(dataframe, half_life_days):
    df = pd.DataFrame()
    for i in range(0,len(dataframe.columns)):
        ti_raw_data = list(dataframe.iloc[:,i])
        half_life = half_life_days
        smoothing_factor = 1 - math.exp(math.log(0.5) / half_life)
        smoothed_values = [ti_raw_data[0]]
        for index in range(1, len(ti_raw_data)):
            previous_smooth_value = smoothed_values[-1]
            new_unsmooth_value = ti_raw_data[index]
            new_smooth_value = ((smoothing_factor * new_unsmooth_value)
                    + ((1 - smoothing_factor) * previous_smooth_value))
            smoothed_values.append(new_smooth_value)
        temp_df = pd.DataFrame(smoothed_values)
        df = pd.concat([df,temp_df],axis=1)
    df.columns = dataframe.columns
    df.index = dataframe.index
    return df

weight_df_ewma = ewma_df(weight_df, half_life_days=90)
weight_df_ewma.plot()
weight_df_diff = weight_df_ewma - weight_df_ewma.shift(1)
weight_df_diff = weight_df_diff.dropna()
weight_df_diff.plot()

style_price = df_asset.iloc[:,6:-1]
rtn_df = pd.DataFrame(np.log(style_price) - np.log(style_price.shift(1)))
rtn_df = rtn_df.dropna()
rtn_df = rtn_df.loc[rtn_df.index >= weight_df_diff.index[0]]

weight_df_diff = weight_df_diff.drop('7-10Y UST', axis=1)


n_of_stocks = 6
n_of_features = 4
n_of_time = 60    # or 40, depends, look back window size
n_of_future = 20  # forward window size for the optimal portfolio

# 1. split train-test
# --------------------------------------------------
# Create sequence data
def make_sequence(x):
    T = n_of_time + n_of_future
    x_seq = np.expand_dims(np.array(x.iloc[0:T, :]), 0)
    
    for i in range(1, len(x) - T + 1):
        d = np.expand_dims(np.array(x.iloc[i:(i+T), :]), 0)
        x_seq = np.concatenate((x_seq, d))
        
    return x_seq

n = int(rtn_df.shape[0] * 0.7)

scaler = StandardScaler() 
scaler.fit(weight_df_diff[:n])    # Only the training data are used to fit the scaler transformation,
weight_train_scaled = scaler.transform(weight_df_diff[:n]) 
weight_test_scaled = scaler.transform(weight_df_diff[n:])  # then the scaler is used to transform the test input data.

weight_scaled = pd.concat([pd.DataFrame(weight_train_scaled),pd.DataFrame(weight_test_scaled)],axis=0)
weight_scaled.index=rtn_df.index

weight_scaled = ewma_df(weight_scaled, half_life_days=2)
weight_scaled.plot()



rtn_train = make_sequence(rtn_df[:n])
weight_train = make_sequence(weight_scaled[:n])
rtn_test = make_sequence(rtn_df[n:])
weight_test =  make_sequence(weight_scaled[n:])

# 2. Split each train / test data based on window sizes
# -------------------------------
# Constructing data for learning/predicting the upcoming period(N_FUTURE periods)
xc_train = np.array([x[:n_of_time] for x in weight_train])
xf_train = np.array([x[-n_of_future:] for x in rtn_train])
                     
xc_test = np.array([x[:n_of_time] for x in weight_test])
xf_test = np.array([x[-n_of_future:] for x in rtn_test])

# 4.Record the index for backtesting purposes
test_date = rtn_df[(n + n_of_time):].index

xf_train = xf_train.astype('float32') * 20.0
xf_test = xf_test.astype('float32') * 20.0

# Shuffle (no need)
xc_train, xf_train = shuffle(xc_train, xf_train)


# Regularization Constraints
GAMMA_CONST = 0.1
REG_CONST = 0.02

# Define Loss Function for the DL Model
# In this particular Network,
# customized loss function has to be used --> max(objective) = min(-objective)
# y_pred = Transfer model ouptut (keras 내부 기능)

def custom_objective(y_true, y_pred):
    W = y_pred      
    xf_rtn = y_true
    W = tf.expand_dims(W, axis = 1)   # W = (None, 1, 50)
    R = tf.expand_dims(tf.reduce_mean(xf_rtn, axis = 1), axis = 2) # R = (None, 50, 1)
    C = tfp.stats.covariance(xf_rtn, sample_axis=1)

    rtn = tf.matmul(W, R)  
    vol = tf.matmul(W, tf.matmul(C, tf.transpose(W, perm = [0, 2, 1]))) * GAMMA_CONST
    regulation = tf.reduce_sum(tf.square(W), axis = -1) * REG_CONST
    objective = rtn - vol - regulation
    
    return -tf.reduce_sum(objective, axis=0)


# Create the DL model
xc_input = Input(batch_shape = (None, n_of_time, n_of_features))
h_lstm = LSTM(50, dropout = 0.5)(xc_input)
y_output = Dense(n_of_stocks, activation='tanh')(h_lstm)  # linear projection

# tanh is used to prevent over-fitting from happening
# just for safety ex : [-3, 0.4, 0.2, +20] --> [-0.995, 0.380, 0.197, 1.0]

# early_stopping, restore_best_weights = True
early_stopping = EarlyStopping(monitor = 'val_loss',  patience = 20, restore_best_weights=True)

y_output = Activation('softmax')(y_output)

model = Model(xc_input, y_output)
model.compile(loss = custom_objective, 
              optimizer = Adam(learning_rate = 1e-5))
model.summary()

hist = model.fit(xc_train, xf_train, epochs=1000, batch_size = 100, 
                 validation_data = (xc_test, xf_test),
                 callbacks=[early_stopping])

# Plot the loss function
plt.figure(figsize=(8, 5))
plt.plot(hist.history['loss'], label='Train loss')
plt.plot(hist.history['val_loss'], label='Validation loss')
plt.legend()
plt.show()



# Backtest (Test Set, Simpler Version)
port_value = [10000]   # portfolio의 초기 value
w_hist_model = []
for i in range(0, xc_test.shape[0], n_of_future):
    x = xc_test[i][np.newaxis,:, :]
    w_lstm = model.predict(x)[0]
    w_hist_model.append(w_lstm)
    
       
    m_rtn = np.sum((xf_test[i]/20), axis = 0)
    port_value.append(port_value[-1] * np.exp(np.dot(w_lstm, m_rtn)))


idx = np.arange(0, len(test_date), n_of_future)
perf_df = pd.DataFrame({'DL model':lstm_value}, 
                       index=test_date[idx])

# Backtest (BM(Market), Simpler Version)
msci_us_df = df_asset.iloc[:,-1]
msci_us_df = np.log(msci_us_df[test_date[0]:]).diff().cumsum()+1
msci_us_df.iloc[0] = 1
msci_us_trans = lstm_value[0] / msci_us_df[0]
perf_df['BM(market)'] = msci_us_df * msci_us_trans
perf_df.plot()

# Backtest (BM(Equal Weight), Simpler Version)
crp_value = [10000]   # CRP의 초기 value
w_crp = np.ones(n_of_stocks) / n_of_stocks

for i in range(0, xc_test.shape[0], n_of_future):
 
    m_rtn = np.sum(xf_test[i]/20, axis = 0)
    crp_value.append(crp_value[-1] * np.exp(np.dot(w_crp, m_rtn)))
perf_df['BM(Equal Weight)'] = crp_value

# Backtest (Transfer Learnign, Rolling Basis, Simpler Version)
mpn_value = [10000]
w_hist = []

for i in range(0, xc_test.shape[0], n_of_future):
    
    print(i)

    x = xc_test[i][np.newaxis,:, :]
    w_prt = model.predict(x)[0]
    w_hist.append(w_prt)

    m_rtn = np.sum(xf_test[i]/20, axis = 0)
    mpn_value.append(mpn_value[-1] * np.exp(np.dot(w_prt, m_rtn)))
  
    xc_new = xc_test[i:(i+n_of_future), :, :]
    xf_new = xf_test[i:(i+n_of_future), :, :]
    
    idx = np.random.randint(0, xc_train.shape[0], 80)

    x = np.vstack([xc_new, xc_train[idx]])
    y = np.vstack([xf_new, xf_train[idx]])
    x, y = shuffle(x, y)
   
    model.fit(x, y, epochs=50, batch_size=10, verbose=0)

    xc_train = np.vstack([xc_train, xc_new])
    xf_train = np.vstack([xf_train, xf_new])


perf_df['additional'] = mpn_value


# plot the final result
fig = plt.figure(figsize=(13, 10))
plt.plot(perf_df.iloc[:,0], "b-", label=perf_df.columns[0])
plt.plot(perf_df.iloc[:,1], "r-", label=perf_df.columns[1])
plt.plot(perf_df.iloc[:,2], "k-", label=perf_df.columns[2])
plt.plot(perf_df.iloc[:,3], "p-", label=perf_df.columns[3])
plt.ylabel("cumulative return(natural log)")
plt.legend(loc="upper left")
plt.show()

# plot the final result - different style
import scienceplots
plt.style.use(['science','no-latex'])
fig, ax = plt.subplots(figsize=(15,8), dpi = 400)
ax.plot(perf_df)
fig.legend(perf_df.columns)
ax.set_xlim([perf_df.index[0], perf_df.index[-1]])
ax.grid()
ax.set_title('DL Backtesting Result')


# save the result

wb = load_workbook('c:/users/user/desktop/Python/기초 자산배분 전략/ML DB/'+'딥러닝 결과_2.xlsx')
sheets = wb.sheetnames
targetsheet = wb[sheets[0]]

# 데이터프레임 입력
rows = dataframe_to_rows(perf_df)

for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
         targetsheet.cell(row=r_idx, column=c_idx, value=value)

# 위치와 파일명 입력
wb.save('c:/users/user/desktop/Python/기초 자산배분 전략/Santa Claus is not Coming to town/' + '딥러닝 결과_3' + '.xlsx')








